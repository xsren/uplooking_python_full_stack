# encoding: utf-8
"""
@author: xsren 
@contact: bestrenxs@gmail.com
@site: xsren.me

@version: 1.0
@license: Apache Licence
@file: pack_multiple_excel_to_one.py
@time: 27/07/2017 5:22 PM

将当前目录下的excel合并成一个新的excel
"""
import os
import time
from copy import copy

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.write_only import WriteOnlyCell


def run():
    wb = Workbook()
    ws_w = wb.active
    fpath = '.'

    i = 1

    files = []
    for fname in os.listdir(fpath):
        if fname.endswith(".xlsx"):
            files.append(fname)

    # 安装最后更改时间排序，一开始用getctime但是mac上有问题（一个文件下载然后复制到别的文件夹的时候，ctime是复制的时间），于是改成getmtime
    files.sort(key=lambda x: os.path.getmtime(x))

    for fname in files:
        print fname
        wb_r = load_workbook(filename=fname)
        ws_r = wb_r.worksheets[0]
        j = 1
        for row in ws_r.rows:

            new_cells = []

            if i == 1:
                for cell in row:
                    new_cell = WriteOnlyCell(ws_w, value=cell.value)
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = copy(cell.number_format)
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)
                    new_cells.append(new_cell)
                if j == 1:
                    ws_w.title = ws_r.title
                    ws_w.column_dimensions = ws_r.column_dimensions
                ws_w.append(new_cells)
            else:
                if j >= 4:
                    for cell in row:
                        new_cell = WriteOnlyCell(ws_w, value=cell.value)
                        if cell.column == "A":
                            new_cell.value = (i - 1) * 10 + int(cell.value)
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)
                        new_cells.append(new_cell)
                    ws_w.append(new_cells)
            j += 1
        i += 1

    today = time.strftime('%Y%m%d', time.localtime(time.time()))
    wb.save('new_file_%s.xlsx' % today)


if __name__ == '__main__':
    run()
